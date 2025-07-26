import openpyxl
from NN_Codes.TrainSegmentationParis import *
from collections import Counter
from RequiredFunctions import *
expNumber=50
paris_Raw_Data_Dir="./data/paris"
DATA_DIR=paris_Raw_Data_Dir+"_out/"

def segmentWithMLP():
    testModel()

def saveSegmentedData(pcd_plants,pcd_back,file,model_name="MLP"):
    if not os.path.exists(DATA_DIR + "6_segmentation_result_background_"+model_name+"/"):
        os.makedirs(DATA_DIR + "6_segmentation_result_background_"+model_name+"/")
    if not os.path.exists(DATA_DIR + "6_segmentation_result_plants_"+model_name+"/"):
        os.makedirs(DATA_DIR + "6_segmentation_result_plants_"+model_name+"/")

    points = np.asarray(pcd_plants.points)
    colors = np.asarray(pcd_plants.colors)
    # Combine point coordinates and colors into a new NumPy array
    plants_data = np.column_stack((points, colors[:, :1]))
    np.savetxt(DATA_DIR + "6_segmentation_result_plants_"+model_name+"/" + file[:-4] + "_plants.txt", plants_data, delimiter=',', fmt='%.2f')

    points = np.asarray(pcd_back.points)
    colors = np.asarray(pcd_back.colors)
    # Combine point coordinates and colors into a new NumPy array
    back_data = np.column_stack((points, colors[:, :1]))
    np.savetxt(DATA_DIR + "6_segmentation_result_background_"+model_name+"/" + file[:-4] + "_background.txt", back_data, delimiter=',', fmt='%.2f')
    print(model_name,"---Segmentation process completed --- check Data/Data2Clean/6_segmentation_result_plants/")

def segmentRansac(distance_threshold = 0.05):
    # Start Segmentation with Pre-Trained model
    input_dir=DATA_DIR + "5_smooted"
    source_directory_list = os.listdir(input_dir)

    # Set the parameters for region growing segmentation.
     # Distance threshold for plane segmentation
    ransac_n = 10  # RANSAC algorithm parameter
    num_iterations = 100  # Number of RANSAC iterations

    for file in source_directory_list:
        df = pd.read_csv(input_dir + "/" + file, header=None)
        point_data = df.to_numpy()
        # Create an Open3D point cloud from the NumPy array
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(point_data[:, :3])  # Set x, y, z coordinates
        pcd.colors = o3d.utility.Vector3dVector(point_data[:, [3,3,3]])  # Set intensity as color

        # Perform plane segmentation to remove ground or large planar surfaces.
        distance_threshold=0.05
        increment=0.05
        while distance_threshold<5:
            inliers, outliers = pcd.segment_plane(distance_threshold, ransac_n, num_iterations)
            pcd_plants = pcd.select_by_index(outliers,invert=True)
            pcd_back = pcd.select_by_index(outliers)

            pcd_plants_t = np.concatenate((np.asarray(pcd_plants.points), np.asarray(pcd_plants.colors)[:,0].reshape(-1,1)), axis=1)
            pcd_back_t = np.concatenate((np.asarray(pcd_back.points), np.asarray(pcd_back.colors)[:,0].reshape(-1,1)), axis=1)

            y_pred = np.ones(len(point_data))
            y_pred[outliers] = 0

            y_data = np.zeros(len(point_data))
            y_data[point_data[:, 4]==1] = 1

            print("file:",file)
            if not os.path.exists(DATA_DIR + "Ransac_result/"):
                os.makedirs(DATA_DIR + "Ransac_result/")

            np.savetxt(DATA_DIR + "Ransac_result/"+file + "_THR_" + str(distance_threshold)+"_train_background.txt", pcd_back_t, fmt="%.1f")
            np.savetxt(DATA_DIR + "Ransac_result/"+file + "_THR_" + str(distance_threshold)+"_train_plants.txt", pcd_plants_t, fmt="%.1f")

            accuracy, precision, recall, f1 = evaluateResults(y_pred, y_data)
            print("-----" + file + "-----")
            write2Excel(DATA_DIR + "Ransac_result/evaluation_metrics.xlsx", file + "_THR_" + str(distance_threshold), accuracy, precision, recall, f1)

            distance_threshold=distance_threshold+increment

def segmentDBScan(normal_radius = 0.1,distance_threshold = 25   ):
    # Start Segmentation with Pre-Trained model
    input_dir=DATA_DIR + "5_smooted"
    source_directory_list = os.listdir(input_dir)

    for file in source_directory_list:
        pcd = o3d.io.read_point_cloud(input_dir + "/" + file)

        X = np.concatenate((np.asarray(pcd.points), np.asarray(pcd.colors) * 255), axis=1)
        pcd = pcd.select_by_index(np.where(X[:, 2] < 200)[0])
        X = X[X[:, 2] < 200]
        X_org = X.copy()

        # Perform plane segmentation to remove ground or large planar surfaces.

        # Perform DBSCAN segmentation
        with o3d.utility.VerbosityContextManager(o3d.utility.VerbosityLevel.Debug):
            labels = pcd.cluster_dbscan(eps=3, min_points=10)

        # Count the points in each cluster
        point_counts = Counter(labels)
        # Find the label of the largest cluster
        largest_cluster_label = max(point_counts, key=point_counts.get)

        labels = np.array(labels)
        largest_cluster_indices = np.where(labels == largest_cluster_label)
        # Create a point cloud with the remaining points.
        pcd_plants = pcd.select_by_index(largest_cluster_indices[0])
        pcd_back = pcd.select_by_index(largest_cluster_indices[0], invert=True)

        plants= np.concatenate((np.asarray(pcd_plants.points), np.asarray(pcd_plants.colors) * 255), axis=1)
        "Inversion of Z coordinate"
        plants[:, 2] = plants[:, 2] * -1

        saveSegmentedData(pcd_plants, pcd_back, file,plants,"DBScan")

def segmentDBScanFromMerged(normal_radius = 0.1,distance_threshold = 25   ):
    # Start Segmentation with Pre-Trained model
    input_dir=DATA_DIR + "3_merged"
    source_directory_list = os.listdir(input_dir)

    for file in source_directory_list:
        pcd = o3d.io.read_point_cloud(input_dir + "/" + file)

        # X = np.concatenate((np.asarray(pcd.points), np.asarray(pcd.colors) * 255), axis=1)
        X=np.asarray(pcd.points)
        pcd = pcd.select_by_index(np.where(X[:, 2] < 200)[0])
        X = X[X[:, 2] < 200]
        X_org = X.copy()

        # Perform plane segmentation to remove ground or large planar surfaces.

        # Perform DBSCAN segmentation
        with o3d.utility.VerbosityContextManager(o3d.utility.VerbosityLevel.Debug):
            labels = pcd.cluster_dbscan(eps=1, min_points=10)

        # Count the points in each cluster
        point_counts = Counter(labels)
        # Find the label of the largest cluster
        largest_cluster_label = max(point_counts, key=point_counts.get)

        labels = np.array(labels)
        largest_cluster_indices = np.where(labels == largest_cluster_label)
        # Create a point cloud with the remaining points.
        pcd_plants = pcd.select_by_index(largest_cluster_indices[0])
        pcd_back = pcd.select_by_index(largest_cluster_indices[0], invert=True)

        plants= np.concatenate((np.asarray(pcd_plants.points), np.asarray(pcd_plants.colors) * 255), axis=1)
        "Inversion of Z coordinate"
        plants[:, 2] = plants[:, 2] * -1

        saveSegmentedData(pcd_plants, pcd_back, file,plants,"DBScanMerged")

def RGBSegmentation(smoothness = 2,cls_size=30):
    # Start Segmentation with Pre-Trained model
    input_dir=DATA_DIR + "5_smooted" # 5_smooted 1_normalized
    source_directory_list = os.listdir(input_dir)

    for file in source_directory_list:
        # for smoothness in range(1,10,1):
        smoothness=0.25
        increment=0.25
        while smoothness<5:
            for curvature in [5,10,20,50]:
                df = pd.read_csv(input_dir + "/" + file, header=None)
                X = df.to_numpy()[:,:3].astype(np.float32)
                X_org = df.to_numpy()[:,:5].astype(np.float32)

                cloud = pcl.PointCloud()
                cloud.from_array(X)

                tree = cloud.make_kdtree()
                segment = cloud.make_RegionGrowing(ksearch=50)
                segment.set_MinClusterSize(30)
                segment.set_MaxClusterSize(10000000)
                segment.set_NumberOfNeighbours(30)
                segment.set_SmoothnessThreshold((smoothness / 180.0) * (3.14))
                segment.set_CurvatureThreshold(curvature)
                segment.set_SearchMethod(tree)
                clusters = segment.Extract()

                index_of_largest = max(range(len(clusters)), key=lambda i: len(clusters[i]))
                solid_index=clusters[index_of_largest]

                pcd_back=X_org[solid_index,:]
                pcd_plants= X_org[~np.isin(np.arange(len(X_org)), solid_index)]

                y_data = np.ones(len(X_org))
                y_data[X_org[:, 4]!=1] = 0 # ground =1
                y_pred = np.ones(len(X_org))
                y_pred[~np.isin(np.arange(len(X_org)), solid_index)]=0

                if not os.path.exists(DATA_DIR + "RGBSegmentation_result/"):
                    os.makedirs(DATA_DIR + "RGBSegmentation_result/")

                np.savetxt(DATA_DIR + "RGBSegmentation_result/"+file+"_SMT_"+str(smoothness)+"_CUR_"+str(curvature)+"_train_background.txt", pcd_back, fmt="%.1f")
                np.savetxt(DATA_DIR + "RGBSegmentation_result/"+file+"_SMT_"+str(smoothness)+"_CUR_"+str(curvature)+"_train_plants.txt", pcd_plants, fmt="%.1f")

                accuracy, precision, recall, f1 = evaluateResults(y_pred, y_data)
                print("-----"+file+"-----")
                write2Excel(DATA_DIR + "RGBSegmentation_result/evaluation_metrics.xlsx",file+"_SMT_"+str(smoothness)+"_CUR_"+str(curvature),accuracy,precision,recall,f1)
                smoothness=smoothness+increment

def write2Excel(excel_file_path,file,accuracy,precision,recall,f1):
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active

        # Define the data for the new row
        data_to_add = [file, round(accuracy, 2), round(precision, 2), round(recall, 2), round(f1, 2)]

        # Find the next available row (assumes that data starts from row 2)
        next_row = worksheet.max_row + 1

        # Add the data to the new row
        for column, value in enumerate(data_to_add, start=1):
            worksheet.cell(row=next_row, column=column, value=value)
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Define the header row
        header = ["file_name", "accuracy", "precision", "recall", "f1"]
        worksheet.append(header)

        # Define the data for the new row
        data_to_add = [file, round(accuracy, 2), round(precision, 2), round(recall, 2), round(f1, 2)]

        # Add the data to the new row
        worksheet.append(data_to_add)

    # Save the new Excel file
    workbook.save(excel_file_path)

if __name__ == "__main__":
    # data.ProcessParisData.preprocessData()
    segmentWithMLP()
    # segmentRansac()
    # segmentDBScanFromMerged()
    # RGBSegmentation()